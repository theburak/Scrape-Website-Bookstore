import os
import requests
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from bs4 import BeautifulSoup
import time
import openpyxl
from openpyxl.drawing.image import Image  # For adding images to Excel

# Scrape All Books "Title", "Price", "Stock Availability"
# from https://books.toscrape.com/

# ChromeDriver settings
service = Service('C:/WebDriver/chromedriver.exe')  # Specify the path to chromedriver
options = webdriver.ChromeOptions()
options.add_argument('--headless')  # Run the browser in the background
driver = webdriver.Chrome(service=service, options=options)

# Create an Excel file
wb = openpyxl.Workbook()
sheet = wb.active
sheet.title = "Books"
sheet.append(["Title", "Price", "Stock Availability", "Cover Image", "Star Rating", "Book URL"])

# Navigate to the website
base_url = "https://books.toscrape.com/catalogue/page-{}.html"

for page in range(1, 51):  # There are 50 pages
    url = base_url.format(page)
    driver.get(url)
    time.sleep(5)  # Wait for the page to load

    # Parse the page source with BeautifulSoup
    soup = BeautifulSoup(driver.page_source, 'html.parser')
    books = soup.find_all('article', class_='product_pod')

    if not books:
        break

    row = 2  # Start adding data from the second row in Excel
    for book in books:
        # Extract book details
        title = book.h3.a['title']
        price = book.find('p', class_='price_color').text
        stock_status = book.find('p', class_='instock availability').text.strip()

        # Extract the book cover image URL
        image_url = "https://books.toscrape.com/" + book.find('img')['src']

        # Extract the book cover image URL and download it
        image_dir = "book_images"
        if not os.path.exists(image_dir):
            os.makedirs(image_dir)  # Create the directory if it doesn't exist
        image_filename = os.path.join(image_dir, title.replace('/', '-') + ".jpg")
        try:
            response = requests.get(image_url, stream=True)
            if response.status_code == 200:
                with open(image_filename, 'wb') as f:
                    f.write(response.content)
        except Exception as e:
            print(f"Failed to download image for {title}: {e}")
            image_filename = None

        # Extract star rating
        star_class = book.find('p', class_='star-rating')['class']
        # Convert star rating class to a number (e.g., "Three" -> 3)
        star_rating = star_class[1]  # Second class contains the rating text
        star_rating_map = {"One": 1, "Two": 2, "Three": 3, "Four": 4, "Five": 5}
        star_rating = star_rating_map.get(star_rating, 0)  # Default to 0 if not found
        # Extract book URL
        book_url = "https://books.toscrape.com/catalogue/" + book.h3.a['href']

        # Write to the Excel file
        sheet.append([title, price, stock_status, image_url, star_rating, book_url])

        # Write text data to the Excel file
        sheet.cell(row=row, column=1, value=title)
        sheet.cell(row=row, column=2, value=price)
        sheet.cell(row=row, column=3, value=stock_status)
        sheet.cell(row=row, column=5, value=star_rating)
        sheet.cell(row=row, column=6, value=book_url)

        # Add the image to Excel if it exists
        if image_filename and os.path.exists(image_filename):
            img = Image(image_filename)
            img.width, img.height = 80, 80  # Resize image for Excel cell
            sheet.add_image(img, f"D{row}")  # Add image to column D

        row += 1

# Close the browser
driver.quit()

# Set row height and column width to 80
for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
    for cell in row:
        sheet.row_dimensions[cell.row].height = 80
        sheet.column_dimensions[cell.column_letter].width = 80

# Save the Excel file
wb.save("books.xlsx")
print("Data has been saved to 'books.xlsx'.")